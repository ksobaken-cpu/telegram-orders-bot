import asyncio
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aiogram import Bot, Dispatcher, F, types
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    FSInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from dotenv import load_dotenv

# 1. Загрузка переменных
load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", 876876919))

# Сброс прокси
os.environ['HTTP_PROXY'] = ''
os.environ['HTTPS_PROXY'] = ''
os.environ['ALL_PROXY'] = ''

bot = Bot(token=TOKEN)
dp = Dispatcher()

EXCEL_FILE = "orders.xlsx"


class OrderForm(StatesGroup):
    category = State()
    details = State()


# Главное меню (всегда зафиксировано внизу)
def get_main_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📝 Оставить заявку")]],
        resize_keyboard=True,
    )


# Клавиатура выбора категорий + кнопка отмены
def get_categories_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🛠 Техподдержка"), KeyboardButton(text="💳 Вопросы по оплате")],
            [KeyboardButton(text="💡 Идея / Предложение"), KeyboardButton(text="❓ Другое")],
            [KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
    )


# Клавиатура только с кнопкой отмены (на этапе ввода текста)
def get_cancel_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="❌ Отмена")]],
        resize_keyboard=True,
    )


# Красивое сохранение в Excel
def save_to_excel(category, user_name, user_id, text):
    try:
        date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        headers = ["Дата и время", "Категория", "Имя", "ID пользователя", "Текст заявки", "Статус"]

        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Заявки"
            ws.append(headers)
        else:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active

        ws.append([date_now, category, user_name, user_id, text, "Новая"])

        # Стилизация
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(vertical="center")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        wb.save(EXCEL_FILE)
        return True
    except PermissionError:
        print(f"ОШИБКА: Файл {EXCEL_FILE} открыт в Excel!")
        return False


# ОБРАБОТЧИК ОТМЕНЫ (срабатывает на команду /cancel или кнопку "❌ Отмена")
@dp.message(Command("cancel"))
@dp.message(F.text == "❌ Отмена")
async def cancel_handler(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is None:
        await message.answer("Нечего отменять.", reply_markup=get_main_keyboard())
        return

    await state.clear()
    await message.answer("🚫 Создание заявки отменено.", reply_markup=get_main_keyboard())


@dp.message(CommandStart())
async def start_cmd(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer(
        f"Привет, {message.from_user.first_name}! 👋\nНажми кнопку ниже, чтобы оформить заявку.",
        reply_markup=get_main_keyboard(),
    )


@dp.message(Command("get_excel"))
async def send_excel(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        await message.answer("⛔ У вас нет прав для выполнения этой команды.")
        return

    if os.path.exists(EXCEL_FILE):
        await message.answer_document(FSInputFile(EXCEL_FILE), caption="📊 Актуальный файл с заявками")
    else:
        await message.answer("⚠️ Файл с заявками пока не создан.")


@dp.message(F.text == "📝 Оставить заявку")
async def start_order(message: types.Message, state: FSMContext):
    await state.set_state(OrderForm.category)
    await message.answer("Выберите тему вашей заявки из меню ниже:", reply_markup=get_categories_keyboard())


@dp.message(OrderForm.category)
async def process_category(message: types.Message, state: FSMContext):
    await state.update_data(chosen_category=message.text)
    await state.set_state(OrderForm.details)

    await message.answer(
        f"Выбрана категория: **{message.text}**\n\nТеперь опишите вашу проблему или вопрос текстом (или нажмите «❌ Отмена»):",
        parse_mode="Markdown",
        reply_markup=get_cancel_keyboard(),
    )


@dp.message(OrderForm.details)
async def process_details(message: types.Message, state: FSMContext):
    data = await state.get_data()
    category = data.get("chosen_category", "Не указана")

    user_name = message.from_user.full_name
    username = f"@{message.from_user.username}" if message.from_user.username else "нет юзернейма"
    user_id = message.from_user.id
    order_text = message.text

    saved = save_to_excel(category, user_name, user_id, order_text)

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="✅ Отметить выполненной", callback_data=f"done_{user_id}")]]
    )

    admin_text = (
        f"🔔 **НОВАЯ ЗАЯВКА!**\n\n"
        f"📂 **Категория:** {category}\n"
        f"👤 **От кого:** {user_name} ({username})\n"
        f"🆔 **ID:** `{user_id}`\n\n"
        f"📝 **Текст:**\n{order_text}\n\n"
        f"📌 **Статус:** 🟡 В обработке"
    )

    try:
        await bot.send_message(chat_id=ADMIN_ID, text=admin_text, parse_mode="Markdown", reply_markup=keyboard)
    except Exception as e:
        print(f"Ошибка отправки админу: {e}")

    if saved:
        await message.answer(
            f"✅ Спасибо, {user_name}! Ваша заявка по теме «{category}» принята.",
            reply_markup=get_main_keyboard(),
        )
    else:
        await message.answer(
            "⚠️ Заявка отправлена менеджеру, но пока не записана в файл.",
            reply_markup=get_main_keyboard(),
        )

    await state.clear()


@dp.callback_query(F.data.startswith("done_"))
async def mark_as_done(callback: types.CallbackQuery):
    target_user_id = callback.data.split("_")[1]
    new_text = callback.message.text.replace("📌 Статус: 🟡 В обработке", "📌 Статус: ✅ Выполнено")

    await callback.message.edit_text(text=f"{new_text}\n\n🎉 *Заявка успешно закрыта!*", parse_mode="Markdown")

    try:
        await bot.send_message(chat_id=target_user_id, text="🎉 Твоя заявка успешно обработана и выполнена!")
    except Exception as e:
        print(f"Не удалось отправить статус пользователю: {e}")

    await callback.answer("Статус обновлен!")


async def main():
    print("Бот запущен и ждёт сообщений...")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())